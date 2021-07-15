import openpyxl

xlfile=openpyxl.Workbook()

default_sheet=xlfile.active

#This is the default sheet which gets created when Excle file is created.
default_sheet.title="Default Sheet"

#This will create new sheet with name "My First Sheet"
new_Sheet1=xlfile.create_sheet("My First Sheet")

#This will change the value in cell A1 to "Open py Excel Sample Data"
new_Sheet1.cell(row=1,column=1).value="Open py Excel Sample Data"

xlfile.save("Openpy_Excel.xlsx")